#!/usr/bin/python2

#Test some functionality of openpyxl module

import openpyxl

wb = openpyxl.load_workbook(filename='src/EVA_Submission_template.V1.1.0.xlsx')
ws = wb['Submitter Details']
ws.cell(row=2, column=1, value='Sichong')
ws['B2'] = 'Peng'
ws['C2'] = '6288885164'
print(ws['A2'].value)
